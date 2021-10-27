from openpyxl import load_workbook

workbook = load_workbook(filename='4A5A港澳台.xlsx')
sheets = workbook.sheetnames
ws = workbook[sheets[2]]  # 5A景区
names = ws['D']  # 景点名称
ids = ws['H']  # 景点ID
urls = ws['G']  # 景点murl

# add_params = {}
#
# for name, jid, url in zip(names[130:131], ids[130:131], urls[130:131]):
#     print(name.value, jid.value)
#
#
# # dd = {'a': 'a', 'b': 'b'}
# # print(list(dd.values()))

import pymysql

pymysql.escape_string()
