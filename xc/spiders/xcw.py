import json
import re

import scrapy
from jsonpath import jsonpath
from openpyxl import load_workbook
from ..settings import logger
from ..items import tdyItem,zbItem

'''
代码功能：景点信息
'''


class XcwSpider(scrapy.Spider):
    name = 'xcw'

    allowed_domains = ['ctrip.com']

    comment_params = {"arg": {"resourceId": 26184, "resourceType": 11, "pageIndex": 1, "pageSize": 10, "sortType": 3,
                              "commentTagId": -1, "collapseType": 1, "channelType": 7, "videoImageSize": "700_392",
                              "starType": 0},
                      "head": {"cid": "09031133417600395476", "ctok": "", "cver": "1.0", "lang": "01", "sid": "8888",
                               "syscode": "09", "auth": None, "extension": [{"name": "protocal", "value": "https"}]},
                      "contentType": "json"}

    source_params = {"arg": {"resourceId": 5078351, "resourceType": 11},
                     "head": {"cid": "09031173417826624428", "ctok": "", "cver": "1.0", "lang": "01", "sid": "8888",
                              "syscode": "09", "auth": None, "extension": [{"name": "protocal", "value": "https"}]},
                     "contentType": "json"}

    detail_info_params = {"poiId": "36552271", "scene": "basic", "head": {"cid": "09031173417826624428", "ctok": "",
                                                                          "cver": "1.0", "lang": "01", "sid": "8888",
                                                                          "syscode": "09", "auth": "",
                                                                          "xsid": "", "extension": []}}

    tdy_api_params = {"useSightExtend": True, "districtId": 1446916, "scene": "basic",
                      "head": {"cid": "09031108218702408075", "ctok": "", "cver": "832.000", "lang": "01",
                               "sid": "8888", "syscode": "09", "auth": "", "xsid": "", "extension": []},
                      "businessId": "1412255"}

    zb_api_params = {"districtId": 1446916, "scene": "basic",
                     "head": {"cid": "09031108218702408075", "ctok": "", "cver": "832.000",
                              "lang": "01", "sid": "8888", "syscode": "09", "auth": "",
                              "xsid": "", "extension": []},
                     "businessId": "1412255"
                     }

    comment_api = "https://m.ctrip.com/restapi/soa2/13444/json/getCommentCollapseList"
    source_api = "https://m.ctrip.com/restapi/soa2/13444/json/getPoiCommentInfoWithHotTag"
    detail_info_api = "https://m.ctrip.com/restapi/soa2/18254/json/getPoiMoreDetail"
    tdy_api = 'https://m.ctrip.com/restapi/soa2/18254/json/GetSightOverview'  # 开园时间、地图、预约  --接口
    zb_api = 'https://m.ctrip.com/restapi/soa2/18254/json/getSightExtendInfo'  # 周边  --接口

    # url = 'https://gs.ctrip.com/html5/you/sight/huangshanscenicarea19/{}.html'
    num = 0

    def start_requests(self):
        workbook = load_workbook(filename='4A5A港澳台.xlsx')
        sheets = workbook.sheetnames
        ws = workbook[sheets[2]]  # 5A景区
        names = ws['D']  # 景点名称
        ids = ws['H']  # 景点ID
        urls = ws['G']  # 景点murl
        for name, jid, url in zip(names[1:], ids[1:], urls[1:]):
            add_params = {}
            add_params['J_name'] = name.value  # 景点名称
            add_params['J_id'] = jid.value  # 景点ID
            j_url = url.value  # 景点url
            logger.info(f'{name.value} {jid.value} 任务发送成功！')
            yield scrapy.Request(url=j_url, callback=self.parse_index_data, cb_kwargs=add_params, dont_filter=True,
                                 errback=self.errback)

    def parse_index_data(self, response, J_name, J_id):
        logger.info(f'{J_name} {J_id} 开始采集~')
        j_kargs = {}
        j_kargs['J_name'] = J_name
        j_kargs['J_id'] = J_id
        j_kargs['level'] = response.xpath('//span[@class="titleTag"]//text()').extract_first()  # 等级 4A 5A
        j_kargs['overallScore'] = response.xpath('//span[@class="commentNum"]/text()').extract_first()  # 总体评分
        j_kargs['label'] = "".join(response.xpath('//div[@class="shortFeatures"]//text()').extract())  # 景区类型  标签

        poiId = (re.search('poiId.*?:(\d+),', response.text, re.S)).group(1)
        districtId = (re.search('districtId.*?:(\d+),', response.text, re.S)).group(1)
        businessId = (re.search('businessId.*?:(\d+),', response.text, re.S)).group(1)
        j_kargs['poiId'] = poiId
        j_kargs['districtId'] = districtId
        j_kargs['businessId'] = businessId

        # self.num += 1
        # logger.info(f'景点信息接口累计请求次数：》》》》》{self.num}《《《《《')

        # 景点主页信息 --已采集
        # self.source_params['arg']['resourceId'] = J_id
        # yield scrapy.Request(self.source_api, callback=self.get_comm_api, method='POST',
        #                      body=json.dumps(self.source_params), cb_kwargs=j_kargs, dont_filter=True, errback=self.errback)

        # 开园时间、地图、预约
        self.tdy_api_params['districtId'] = districtId
        self.tdy_api_params['businessId'] = businessId
        yield scrapy.Request(self.tdy_api, callback=self.get_tdy_info, method='POST',
                             body=json.dumps(self.tdy_api_params), cb_kwargs=j_kargs, dont_filter=True,
                             errback=self.errback)

        # 周边
        self.zb_api_params['districtId'] = districtId
        self.zb_api_params['businessId'] = businessId
        yield scrapy.Request(self.zb_api, callback=self.get_zb_info, method='POST',
                             body=json.dumps(self.zb_api_params), cb_kwargs=j_kargs, dont_filter=True,
                             errback=self.errback)

        # 发送景区详情页信息任务
        # da_info = {}
        # da_info['J_name'] = J_name
        # da_info['J_id'] = J_id
        # yield scrapy.Request(self.detail_info_api, callback=self.get_detail_info, method='POST',
        #                      body=json.dumps(self.detail_info_params), cb_kwargs=da_info, dont_filter=True)

    def get_comm_api(self, response, J_name, J_id, level, overallScore, label, poiId):
        json_data = json.loads(response.text)
        totalCount = jsonpath(json_data, '$..poiInfo.commentCount')[0]  # 评论数量
        js, qw, xjb = jsonpath(json_data, '$..scores..score')  # 分维度评分
        # print("ID ", J_id)
        # print("名称 ", J_name)
        # print("等级 ", level)
        # print("总评分 ", overallScore)
        # print("标签 ", label)
        # print("评论数 ", totalCount)
        # print(js, qw, xjb)
        # print("===" * 45)

        logger.info(f'{J_name} {J_id} 采集完成~')
        logger.info('---' * 20)

        item = {}
        item['J_id'] = J_id
        item['J_name'] = J_name
        item['level'] = level
        item['overallScore'] = overallScore
        item['label'] = label
        item['totalCount'] = totalCount
        item['js'] = js
        item['qw'] = qw
        item['xjb'] = xjb
        yield item

    def get_tdy_info(self, response, J_name, J_id, level, overallScore, label, poiId):
        try:
            json_data = json.loads(response.text)
            openTime = jsonpath(json_data, '$..openTime')
            address = jsonpath(json_data, '$..address')
            noticeAppointment = jsonpath(json_data, '$..noticeAppointment')

            item = tdyItem()
            item['openTime'] = openTime if openTime else ''
            item['address'] = address if address else ''
            item['noticeAppointment'] = noticeAppointment if noticeAppointment else ''
            yield item
        except Exception as e:
            logger.error(f'----------{J_name} - {J_id} T D Y 信息解析异常，{e}')

    def get_zb_info(self, response, J_name, J_id, level, overallScore, label, poiId):
        try:
            json_data = json.loads(response.text)
            trafficDesc = jsonpath(json_data, '$..trafficDesc')

            item = zbItem()
            item['trafficDesc'] = trafficDesc if trafficDesc else ''
            yield item
        except Exception as e:
            logger.error(f'----------{J_name} - {J_id} 周边信息解析异常，{e}')

    def get_detail_info(self, response, J_name, J_id, level, overallScore, label, poiId):
        pass

    def errback(self, failure):
        response = failure.value.response
        with open('error_url.txt', 'a')as f:
            f.write(failure)
            f.write(response.url)
            # f.write(response.body)
